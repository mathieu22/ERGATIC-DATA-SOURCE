from flask import Blueprint, request, jsonify, Response, send_file
from sqlalchemy import or_
from app.models import UniteLegale, Etablissement
from app import db
from app.utils.geo import lambert93_to_gps, format_gps_link
import csv
import io
from datetime import datetime

export_bp = Blueprint('export', __name__)


# ============================================
# EXPORT EXCEL - Entreprise complète
# ============================================

@export_bp.route('/entreprise/<siren>/excel')
def export_entreprise_excel(siren):
    """
    Export Excel complet d'une entreprise
    - Feuille 1 : Unité légale (infos entreprise)
    - Feuille 2 : Établissements (tous)
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        return jsonify({'error': 'pandas/openpyxl non installé'}), 500

    if not siren.isdigit() or len(siren) != 9:
        return jsonify({'error': 'SIREN invalide'}), 400

    entreprise = db.session.query(UniteLegale).get(siren)
    if not entreprise:
        return jsonify({'error': 'Entreprise non trouvée'}), 404

    # Récupérer les établissements
    etablissements = db.session.query(Etablissement).filter_by(
        siren=siren
    ).order_by(
        Etablissement.etablissement_siege.desc(),
        Etablissement.etat_administratif.asc()
    ).all()

    # Créer le workbook
    wb = Workbook()

    # ========== FEUILLE 1 : UNITÉ LÉGALE ==========
    ws_ul = wb.active
    ws_ul.title = "Unité Légale"

    # Style en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Données unité légale
    ul_data = [
        ("SIREN", entreprise.siren),
        ("Dénomination", entreprise.denomination or ''),
        ("Dénomination usuelle 1", entreprise.denomination_usuelle_1 or ''),
        ("Dénomination usuelle 2", entreprise.denomination_usuelle_2 or ''),
        ("Dénomination usuelle 3", entreprise.denomination_usuelle_3 or ''),
        ("Sigle", entreprise.sigle or ''),
        ("Nom (personne physique)", entreprise.nom or ''),
        ("Prénom", entreprise.prenom_1 or ''),
        ("Prénom usuel", entreprise.prenom_usuel or ''),
        ("Sexe", entreprise.sexe or ''),
        ("Catégorie juridique", entreprise.categorie_juridique or ''),
        ("Activité principale (NAF)", entreprise.activite_principale or ''),
        ("Nomenclature activité", entreprise.nomenclature_activite or ''),
        ("Catégorie entreprise", entreprise.categorie_entreprise or ''),
        ("Tranche effectifs", entreprise.tranche_effectifs or ''),
        ("Année effectifs", entreprise.annee_effectifs or ''),
        ("État administratif", "Actif" if entreprise.etat_administratif == 'A' else "Cessé"),
        ("Économie sociale solidaire", "Oui" if entreprise.economie_sociale_solidaire == 'O' else "Non"),
        ("Société à mission", "Oui" if entreprise.societe_mission == 'O' else "Non"),
        ("Caractère employeur", "Oui" if entreprise.caractere_employeur == 'O' else "Non"),
        ("Date création", entreprise.date_creation.strftime('%d/%m/%Y') if entreprise.date_creation else ''),
        ("NIC siège", entreprise.nic_siege or ''),
        ("SIRET siège", entreprise.siret_siege or ''),
        ("Identifiant association", entreprise.identifiant_association or ''),
        ("Statut diffusion", entreprise.statut_diffusion or ''),
        ("Date dernier traitement", entreprise.date_dernier_traitement.strftime('%d/%m/%Y %H:%M') if entreprise.date_dernier_traitement else ''),
    ]

    # Écrire les données
    for row_idx, (label, value) in enumerate(ul_data, start=1):
        ws_ul.cell(row=row_idx, column=1, value=label)
        ws_ul.cell(row=row_idx, column=1).font = Font(bold=True)
        ws_ul.cell(row=row_idx, column=1).fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        ws_ul.cell(row=row_idx, column=2, value=value)

    ws_ul.column_dimensions['A'].width = 30
    ws_ul.column_dimensions['B'].width = 50

    # ========== FEUILLE 2 : ÉTABLISSEMENTS ==========
    ws_etab = wb.create_sheet("Établissements")

    # En-têtes
    headers = [
        "SIRET", "NIC", "Siège", "État", "Dénomination usuelle",
        "Enseigne 1", "Enseigne 2", "Enseigne 3",
        "N° voie", "Type voie", "Libellé voie", "Complément adresse",
        "Code postal", "Commune", "Code commune",
        "Code CEDEX", "Libellé CEDEX",
        "Pays étranger", "Commune étranger",
        "Activité principale (NAF)", "Tranche effectifs", "Année effectifs",
        "Caractère employeur", "Date création", "Date dernier traitement",
        "Latitude GPS", "Longitude GPS", "Lien Google Maps",
        "Lambert X", "Lambert Y"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_etab.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Données établissements
    for row_idx, etab in enumerate(etablissements, start=2):
        # Conversion coordonnées GPS
        lat, lon = lambert93_to_gps(etab.coordonnee_lambert_x, etab.coordonnee_lambert_y)
        gps_link = format_gps_link(lat, lon)

        data = [
            etab.siret,
            etab.nic,
            "Oui" if etab.etablissement_siege else "Non",
            "Actif" if etab.etat_administratif == 'A' else "Fermé",
            etab.denomination_usuelle or '',
            etab.enseigne_1 or '',
            etab.enseigne_2 or '',
            etab.enseigne_3 or '',
            etab.numero_voie or '',
            etab.type_voie or '',
            etab.libelle_voie or '',
            etab.complement_adresse or '',
            etab.code_postal or '',
            etab.libelle_commune or '',
            etab.code_commune or '',
            etab.code_cedex or '',
            etab.libelle_cedex or '',
            etab.libelle_pays_etranger or '',
            etab.libelle_commune_etranger or '',
            etab.activite_principale or '',
            etab.tranche_effectifs or '',
            etab.annee_effectifs or '',
            "Oui" if etab.caractere_employeur == 'O' else "Non",
            etab.date_creation.strftime('%d/%m/%Y') if etab.date_creation else '',
            etab.date_dernier_traitement.strftime('%d/%m/%Y %H:%M') if etab.date_dernier_traitement else '',
            lat or '',
            lon or '',
            gps_link or '',
            float(etab.coordonnee_lambert_x) if etab.coordonnee_lambert_x else '',
            float(etab.coordonnee_lambert_y) if etab.coordonnee_lambert_y else '',
        ]

        for col_idx, value in enumerate(data, start=1):
            ws_etab.cell(row=row_idx, column=col_idx, value=value)

    # Ajuster largeur colonnes
    for col_idx, header in enumerate(headers, start=1):
        ws_etab.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = max(len(header) + 2, 12)

    # Sauvegarder dans un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"entreprise_{siren}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================
# EXPORT CSV - Rétrocompatibilité
# ============================================

@export_bp.route('/csv', methods=['POST'])
def export_csv():
    """Export des résultats en CSV"""
    data = request.get_json()
    sirens = data.get('sirens', [])

    if not sirens:
        return jsonify({'error': 'Liste de SIREN requise'}), 400

    sirens = sirens[:10000]

    entreprises = db.session.query(UniteLegale).filter(
        UniteLegale.siren.in_(sirens)
    ).all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quotechar='"')

    writer.writerow([
        'SIREN', 'Dénomination', 'Sigle', 'Catégorie Juridique',
        'Activité Principale', 'Catégorie Entreprise', 'Tranche Effectifs',
        'État', 'Date Création', 'SIRET Siège', 'Adresse Siège',
        'Code Postal', 'Ville', 'Latitude', 'Longitude'
    ])

    for e in entreprises:
        siege = db.session.query(Etablissement).filter_by(
            siren=e.siren,
            etablissement_siege=True
        ).first()

        lat, lon = (None, None)
        if siege:
            lat, lon = lambert93_to_gps(siege.coordonnee_lambert_x, siege.coordonnee_lambert_y)

        writer.writerow([
            e.siren,
            e.nom_complet,
            e.sigle or '',
            e.categorie_juridique or '',
            e.activite_principale or '',
            e.categorie_entreprise or '',
            e.tranche_effectifs or '',
            'Actif' if e.est_active else 'Cessé',
            e.date_creation.strftime('%d/%m/%Y') if e.date_creation else '',
            siege.siret if siege else '',
            siege.adresse_ligne if siege else '',
            siege.code_postal if siege else '',
            siege.libelle_commune if siege else '',
            lat or '',
            lon or ''
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename=export_entreprises_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        }
    )


@export_bp.route('/etablissements/<siren>/csv')
def export_etablissements_csv(siren):
    """Export des établissements d'une entreprise en CSV avec GPS"""
    if not siren.isdigit() or len(siren) != 9:
        return jsonify({'error': 'SIREN invalide'}), 400

    entreprise = db.session.query(UniteLegale).get(siren)
    if not entreprise:
        return jsonify({'error': 'Entreprise non trouvée'}), 404

    etablissements = db.session.query(Etablissement).filter_by(
        siren=siren
    ).order_by(
        Etablissement.etablissement_siege.desc(),
        Etablissement.etat_administratif.asc()
    ).all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quotechar='"')

    writer.writerow([
        'SIRET', 'NIC', 'Est Siège', 'Dénomination', 'Enseigne',
        'État', 'Activité Principale', 'Adresse', 'Code Postal',
        'Ville', 'Date Création', 'Latitude', 'Longitude', 'Google Maps'
    ])

    for e in etablissements:
        lat, lon = lambert93_to_gps(e.coordonnee_lambert_x, e.coordonnee_lambert_y)
        gps_link = format_gps_link(lat, lon)

        writer.writerow([
            e.siret,
            e.nic,
            'Oui' if e.etablissement_siege else 'Non',
            e.denomination_usuelle or '',
            e.enseigne_1 or '',
            'Actif' if e.est_actif else 'Fermé',
            e.activite_principale or '',
            e.adresse_ligne,
            e.code_postal or '',
            e.libelle_commune or '',
            e.date_creation.strftime('%d/%m/%Y') if e.date_creation else '',
            lat or '',
            lon or '',
            gps_link or ''
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename=etablissements_{siren}_{datetime.now().strftime("%Y%m%d")}.csv'
        }
    )


@export_bp.route('/search/excel')
def export_search_excel():
    """Export des résultats de recherche en Excel avec toutes les infos"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl non installé'}), 500

    q = request.args.get('q', '').strip()
    code_postal = request.args.get('code_postal', '').strip()
    ville = request.args.get('ville', '').strip()
    activite = request.args.get('activite', '').strip()
    categorie = request.args.get('categorie', '').strip()
    etat = request.args.get('etat', '').strip()

    query = db.session.query(UniteLegale)

    if q:
        search_term = f"%{q}%"
        query = query.filter(
            or_(
                UniteLegale.denomination.ilike(search_term),
                UniteLegale.sigle.ilike(search_term),
                UniteLegale.nom.ilike(search_term),
                UniteLegale.siren.like(search_term)
            )
        )

    if activite:
        query = query.filter(UniteLegale.activite_principale.like(f"{activite}%"))
    if categorie:
        query = query.filter(UniteLegale.categorie_entreprise == categorie)
    if etat:
        query = query.filter(UniteLegale.etat_administratif == etat)

    if code_postal or ville:
        subquery = db.session.query(Etablissement.siren).filter(
            Etablissement.etablissement_siege == True
        )
        if code_postal:
            subquery = subquery.filter(Etablissement.code_postal.like(f"{code_postal}%"))
        if ville:
            subquery = subquery.filter(Etablissement.libelle_commune.ilike(f"%{ville}%"))
        query = query.filter(UniteLegale.siren.in_(subquery))

    entreprises = query.limit(10000).all()

    if not entreprises:
        return jsonify({'error': 'Aucun résultat à exporter'}), 404

    # Créer le workbook
    wb = Workbook()

    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    # ========== FEUILLE 1 : ENTREPRISES ==========
    ws_entreprises = wb.active
    ws_entreprises.title = "Entreprises"

    headers_ul = [
        "SIREN", "Dénomination", "Sigle", "Nom", "Prénom",
        "Catégorie juridique", "Activité principale (NAF)", "Catégorie entreprise",
        "Tranche effectifs", "État", "Date création",
        "SIRET Siège", "Adresse siège", "Code postal", "Ville",
        "Latitude GPS", "Longitude GPS", "Google Maps"
    ]

    for col_idx, header in enumerate(headers_ul, start=1):
        cell = ws_entreprises.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, e in enumerate(entreprises, start=2):
        siege = db.session.query(Etablissement).filter_by(
            siren=e.siren,
            etablissement_siege=True
        ).first()

        lat, lon = (None, None)
        gps_link = None
        if siege:
            lat, lon = lambert93_to_gps(siege.coordonnee_lambert_x, siege.coordonnee_lambert_y)
            gps_link = format_gps_link(lat, lon)

        data = [
            e.siren,
            e.denomination or '',
            e.sigle or '',
            e.nom or '',
            e.prenom_1 or '',
            e.categorie_juridique or '',
            e.activite_principale or '',
            e.categorie_entreprise or '',
            e.tranche_effectifs or '',
            'Actif' if e.etat_administratif == 'A' else 'Cessé',
            e.date_creation.strftime('%d/%m/%Y') if e.date_creation else '',
            siege.siret if siege else '',
            siege.adresse_ligne if siege else '',
            siege.code_postal if siege else '',
            siege.libelle_commune if siege else '',
            lat or '',
            lon or '',
            gps_link or ''
        ]

        for col_idx, value in enumerate(data, start=1):
            ws_entreprises.cell(row=row_idx, column=col_idx, value=value)

    # Ajuster largeurs
    for col_idx, header in enumerate(headers_ul, start=1):
        ws_entreprises.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = max(len(header) + 2, 15)

    # ========== FEUILLE 2 : ÉTABLISSEMENTS ==========
    ws_etab = wb.create_sheet("Établissements")

    headers_etab = [
        "SIREN", "SIRET", "NIC", "Siège", "État",
        "Dénomination", "Enseigne",
        "N° voie", "Type voie", "Libellé voie",
        "Code postal", "Commune",
        "Activité principale (NAF)", "Tranche effectifs",
        "Date création",
        "Latitude GPS", "Longitude GPS", "Google Maps"
    ]

    for col_idx, header in enumerate(headers_etab, start=1):
        cell = ws_etab.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Récupérer tous les établissements des entreprises trouvées
    sirens = [e.siren for e in entreprises]
    etablissements = db.session.query(Etablissement).filter(
        Etablissement.siren.in_(sirens)
    ).order_by(
        Etablissement.siren,
        Etablissement.etablissement_siege.desc()
    ).all()

    for row_idx, etab in enumerate(etablissements, start=2):
        lat, lon = lambert93_to_gps(etab.coordonnee_lambert_x, etab.coordonnee_lambert_y)
        gps_link = format_gps_link(lat, lon)

        data = [
            etab.siren,
            etab.siret,
            etab.nic,
            "Oui" if etab.etablissement_siege else "Non",
            "Actif" if etab.etat_administratif == 'A' else "Fermé",
            etab.denomination_usuelle or '',
            etab.enseigne_1 or '',
            etab.numero_voie or '',
            etab.type_voie or '',
            etab.libelle_voie or '',
            etab.code_postal or '',
            etab.libelle_commune or '',
            etab.activite_principale or '',
            etab.tranche_effectifs or '',
            etab.date_creation.strftime('%d/%m/%Y') if etab.date_creation else '',
            lat or '',
            lon or '',
            gps_link or ''
        ]

        for col_idx, value in enumerate(data, start=1):
            ws_etab.cell(row=row_idx, column=col_idx, value=value)

    # Ajuster largeurs
    for col_idx, header in enumerate(headers_etab, start=1):
        ws_etab.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = max(len(header) + 2, 12)

    # Sauvegarder
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"recherche_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@export_bp.route('/search/csv')
def export_search_csv():
    """Export des résultats de recherche en CSV avec GPS"""
    q = request.args.get('q', '').strip()
    code_postal = request.args.get('code_postal', '').strip()
    ville = request.args.get('ville', '').strip()
    activite = request.args.get('activite', '').strip()
    categorie = request.args.get('categorie', '').strip()
    etat = request.args.get('etat', '').strip()

    query = db.session.query(UniteLegale)

    if q:
        search_term = f"%{q}%"
        query = query.filter(
            or_(
                UniteLegale.denomination.ilike(search_term),
                UniteLegale.sigle.ilike(search_term),
                UniteLegale.nom.ilike(search_term),
                UniteLegale.siren.like(search_term)
            )
        )

    if activite:
        query = query.filter(UniteLegale.activite_principale.like(f"{activite}%"))
    if categorie:
        query = query.filter(UniteLegale.categorie_entreprise == categorie)
    if etat:
        query = query.filter(UniteLegale.etat_administratif == etat)

    if code_postal or ville:
        subquery = db.session.query(Etablissement.siren).filter(
            Etablissement.etablissement_siege == True
        )
        if code_postal:
            subquery = subquery.filter(Etablissement.code_postal.like(f"{code_postal}%"))
        if ville:
            subquery = subquery.filter(Etablissement.libelle_commune.ilike(f"%{ville}%"))
        query = query.filter(UniteLegale.siren.in_(subquery))

    entreprises = query.limit(10000).all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quotechar='"')

    writer.writerow([
        'SIREN', 'Dénomination', 'Sigle', 'Catégorie Juridique',
        'Activité Principale', 'Catégorie Entreprise', 'Tranche Effectifs',
        'État', 'Date Création', 'SIRET Siège', 'Adresse Siège',
        'Code Postal', 'Ville', 'Latitude', 'Longitude', 'Google Maps'
    ])

    for e in entreprises:
        siege = db.session.query(Etablissement).filter_by(
            siren=e.siren,
            etablissement_siege=True
        ).first()

        lat, lon = (None, None)
        gps_link = None
        if siege:
            lat, lon = lambert93_to_gps(siege.coordonnee_lambert_x, siege.coordonnee_lambert_y)
            gps_link = format_gps_link(lat, lon)

        writer.writerow([
            e.siren,
            e.nom_complet,
            e.sigle or '',
            e.categorie_juridique or '',
            e.activite_principale or '',
            e.categorie_entreprise or '',
            e.tranche_effectifs or '',
            'Actif' if e.est_active else 'Cessé',
            e.date_creation.strftime('%d/%m/%Y') if e.date_creation else '',
            siege.siret if siege else '',
            siege.adresse_ligne if siege else '',
            siege.code_postal if siege else '',
            siege.libelle_commune if siege else '',
            lat or '',
            lon or '',
            gps_link or ''
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename=recherche_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        }
    )
